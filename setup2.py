import streamlit as st
import os
from datetime import datetime
from PIL import Image
import io
import streamlit.components.v1 as components
from openpyxl import Workbook, load_workbook

st.set_page_config(layout="wide")

# =======================
# Config
# =======================
IMG_DIR = "images"
SAVE_PATH = "scores.xlsx"
VALID_EXTS = (".png", ".jpg", ".jpeg", ".bmp", ".tiff", ".webp")
LABELS = {1: "Poor", 2: "Fair", 3: "Good", 4: "Very Good", 5: "Excellent"}

# =======================
# Fast helpers
# =======================
@st.cache_data(show_spinner=False)
def list_images(img_dir: str):
    return sorted(
        f for f in os.listdir(img_dir)
        if f.lower().endswith(VALID_EXTS) and not f.startswith(".")
    )

@st.cache_data(show_spinner=False)
def load_image_bytes(img_path: str, max_side: int = 1800) -> bytes:
    with Image.open(img_path) as im:
        im = im.convert("RGB")
        im.thumbnail((max_side, max_side), Image.Resampling.LANCZOS)
        buf = io.BytesIO()
        im.save(buf, format="JPEG", quality=90, optimize=True)
        return buf.getvalue()

def append_to_excel_fast(path: str, row: dict):
    headers = ["image", "score", "label", "time"]
    values = [row.get(h, "") for h in headers]

    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "scores"
        ws.append(headers)
        ws.append(values)
        wb.save(path)
        return

    wb = load_workbook(path)
    ws = wb.active
    ws.append(values)
    wb.save(path)

# =======================
# State
# =======================
img_list = list_images(IMG_DIR)

if "idx" not in st.session_state:
    st.session_state.idx = 0

if st.session_state.idx >= len(img_list):
    st.success("所有图像已评分完成！")
    st.stop()

radio_key = f"score_choice_{st.session_state.idx}"

img_name = img_list[st.session_state.idx]
img_path = os.path.join(IMG_DIR, img_name)

# 预加载下一张
if st.session_state.idx + 1 < len(img_list):
    _ = load_image_bytes(
        os.path.join(IMG_DIR, img_list[st.session_state.idx + 1]),
        max_side=1800
    )

# =======================
# CSS
# =======================
st.markdown(
    """
    <style>
    .block-container{
        padding-left: 1.2rem;
        padding-right: 1.2rem;
        max-width: 1800px;
        margin: 0 auto;
    }
    .img-wrap img{
        width: 100% !important;
        height: auto !important;
        max-height: 90vh !important;
        object-fit: contain !important;
        display: block;
        margin: 0 auto;
    }
    .panel{
        position: sticky;
        top: 1rem;
        border: 1px solid rgba(0,0,0,0.10);
        border-radius: 16px;
        padding: 18px;
        background: white;
        box-shadow: 0 8px 24px rgba(0,0,0,0.08);
    }

    /* Radio buttons */
    div[role="radiogroup"] > label{
        margin: 0.18rem 0 !important;
        padding: 0.12rem 0 !important;
    }
    div[role="radiogroup"] input[type="radio"]{
        transform: scale(1.4);
        margin-right: 0.8rem !important;
        accent-color: #e74c3c;
        cursor: pointer;
    }
    div[role="radiogroup"] label span{
        font-size: 18px !important;
        font-weight: 850 !important;
        line-height: 1.15 !important;
    }

    /* Next button — BIG & CLEAR */
    div.stButton > button{
        width: 100%;
        padding: 1.0rem 7.3rem;   /* 👈 关键：高度明显增加 */
        border-radius: 16px;
        font-weight: 900;
        font-size: 18px;
        letter-spacing: 0.3px;
        background-color: #111827;
        color: #ffffff;
        border: 2px solid #111827;
        transition: background-color 0.12s ease-out,
                    box-shadow 0.12s ease-out,
                    transform 0.05s ease-out;
    }
    div.stButton > button:hover:enabled{
        background-color: #000000;
        box-shadow: 0 8px 20px rgba(0,0,0,0.28);
        transform: translateY(-1px);
    }
    div.stButton > button:active:enabled{
        transform: translateY(0px);
        box-shadow: 0 4px 10px rgba(0,0,0,0.25);
    }
    div.stButton > button:disabled{
        background-color: #e5e7eb;
        color: #9ca3af;
        border: 2px solid #d1d5db;
        box-shadow: none;
        cursor: not-allowed;
    }

    .hint{
        font-size: 12px;
        opacity: 0.65;
        margin-top: 6px;
        margin-bottom: 12px;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# =======================
# Layout
# =======================
left, right = st.columns([3.6, 1.4], gap="large")

with left:
    st.markdown('<div class="img-wrap">', unsafe_allow_html=True)
    st.image(
        load_image_bytes(img_path, max_side=1800),
        caption=img_name,
        width="stretch"
    )
    st.markdown("</div>", unsafe_allow_html=True)

with right:
    st.markdown('<div class="panel">', unsafe_allow_html=True)
    st.markdown("### Rate image quality")
    st.markdown(
        '<div class="hint">No back. Select one score, then Next. (→ / Enter)</div>',
        unsafe_allow_html=True
    )

    options = [5, 4, 3, 2, 1]
    score = st.radio(
        label="",
        options=options,
        index=None,
        key=radio_key,
        format_func=lambda x: f"{x}  —  {LABELS[x]}",
        label_visibility="collapsed",
    )

    st.markdown('<div id="next_btn_anchor">', unsafe_allow_html=True)
    next_clicked = st.button("Next", disabled=(score is None))
    st.markdown("</div>", unsafe_allow_html=True)

    components.html(
        """
        <script>
        (function() {
          if (window.__iqa_keybind_final_big__) return;
          window.__iqa_keybind_final_big__ = true;

          function findNextButton(doc) {
            const anchor = doc.querySelector('#next_btn_anchor');
            if (!anchor) return null;
            return anchor.querySelector('button');
          }

          function clickNext() {
            let btn = findNextButton(document);
            if (!btn && window.parent && window.parent.document) {
              btn = findNextButton(window.parent.document);
            }
            if (btn && !btn.disabled) btn.click();
          }

          const handler = function(e) {
            const tag = (e.target && e.target.tagName) ? e.target.tagName.toLowerCase() : "";
            if (tag === "input" || tag === "textarea") return;

            if (e.key === "ArrowRight" || e.key === "Enter") {
              e.preventDefault();
              clickNext();
            }
            if (e.key === "ArrowLeft") {
              e.preventDefault();
            }
          };

          document.addEventListener('keydown', handler, { passive: false });
          if (window.parent && window.parent.document) {
            window.parent.document.addEventListener('keydown', handler, { passive: false });
          }
        })();
        </script>
        """,
        height=0,
    )

    st.markdown("</div>", unsafe_allow_html=True)

# =======================
# Save & Next
# =======================
if next_clicked:
    append_to_excel_fast(
        SAVE_PATH,
        {
            "image": img_name,
            "score": int(score),
            "label": LABELS[int(score)],
            "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }
    )
    st.session_state.idx += 1
    st.rerun()
